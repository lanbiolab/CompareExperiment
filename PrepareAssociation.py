'''
@Author: DongYi
@Description: 这是为将circRNA_cancer
circRNA_miRNA
miRNA_cancer的表格读取出来
20201103 增加了对csv文件的处理
'''

from openpyxl import load_workbook
import numpy as np
import pandas as pd
import h5py
import csv

# # 读取circRNA_cancer_association
#
# circRNA_cancer_association_table = load_workbook('./Data/miRNA_cancer/miRNA_cancer_association.xlsx')
#
# circ_cancer_worksheets = circRNA_cancer_association_table.worksheets
#
# circ_cancer_list = []
# for circan_sheet in circ_cancer_worksheets:
#     rows = circan_sheet.rows
#     for row in rows:
#         row_val = [col.value.lower() for col in row]
#         circ_cancer_list.append(row_val)
#
# # 读取circRNA_list
#
# circRNA_table = load_workbook('./Data/miRNA_cancer/miRNA_list.xlsx')
#
# circ_worksheets = circRNA_table.worksheets
#
# circ_list = []
# for circ_sheet in circ_worksheets:
#     rows = circ_sheet.rows
#     for row in rows:
#         row_val = [col.value.lower() for col in row]
#         circ_list += row_val
#
# # 读取cancer_list
#
# cancer_table = load_workbook('./Data/miRNA_cancer/cancer_list.xlsx')
# cancer_worksheets = cancer_table.worksheets
#
# canc_list =[]
# for canc_sheet in cancer_worksheets:
#     rows = canc_sheet.rows
#     for row in rows:
#         row_val = [col.value.lower() for col in row]
#         canc_list += row_val
#
# # 构造关系矩阵
# circRNA_cancer_association = np.zeros((len(circ_list), len(canc_list)))
#
# for i in range(len(circ_cancer_list)):
#     circRNA_name = circ_cancer_list[i][0]
#     cancer_name = circ_cancer_list[i][1]
#     circRNA_index = circ_list.index(circRNA_name)
#     cancer_index = canc_list.index(cancer_name)
#     circRNA_cancer_association[circRNA_index, cancer_index] = 1
#
# # 存储入h5文件
# with h5py.File('./Data/miRNA_cancer/miRNA_cancer.h5') as f:
#     f['infor'] = circRNA_cancer_association
#
# # 把已经调整好的矩阵输出到一个csv文件中
# circ_can_data = pd.DataFrame(circRNA_cancer_association, index=circ_list, columns=canc_list)
# circ_can_data.to_csv('./Data/miRNA_cancer/miRNA_cancer.csv')

#===============================================================================================================#
'''
读取csv数据文件
'''

circ_dis_list = []
with open('./Data/circRNA_disease_from_circRNADisease/association.csv', 'r') as f:
    reader = csv.reader(f)
    for row in reader: # 把row中string 转为 int
        row = [int(num) for num in row]
        circ_dis_list.append(np.array(row))
circ_dis_array = np.array(circ_dis_list)

with h5py.File('./Data/circRNA_disease_from_circRNADisease/association.h5') as hf:
    hf['infor'] = circ_dis_array



